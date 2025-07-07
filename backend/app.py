import logging
import requests
from bs4 import BeautifulSoup
import re
import json
import os
from datetime import datetime
from flask import Flask, request, jsonify
from flask_cors import CORS
import google.generativeai as genai
from dotenv import load_dotenv
from openpyxl import load_workbook

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Create logs directory if it doesn't exist (in root directory)
LOGS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'logs')
if not os.path.exists(LOGS_DIR):
    os.makedirs(LOGS_DIR)

def log_api_request(endpoint: str, zip_code: str, response_data: dict, extra_data: dict = None):
    """Log API requests and responses to separate files"""
    timestamp = datetime.now().isoformat()

    # Create log entry
    log_entry = {
        'timestamp': timestamp,
        'endpoint': endpoint,
        'zip_code': zip_code,
        'response_data': response_data,
        'extra_data': extra_data or {}
    }

    # Log to main API log file
    api_log_file = os.path.join(LOGS_DIR, 'api_requests.jsonl')
    with open(api_log_file, 'a', encoding='utf-8') as f:
        f.write(json.dumps(log_entry) + '\n')

    # Log to endpoint-specific file
    endpoint_name = endpoint.replace('/', '_').replace('-', '_')
    endpoint_log_file = os.path.join(LOGS_DIR, f'{endpoint_name}_data.jsonl')
    with open(endpoint_log_file, 'a', encoding='utf-8') as f:
        f.write(json.dumps(log_entry) + '\n')

    # Log extra data to separate file if present
    if extra_data:
        extra_log_file = os.path.join(LOGS_DIR, f'{endpoint_name}_extra_data.jsonl')
        extra_entry = {
            'timestamp': timestamp,
            'zip_code': zip_code,
            'extra_data': extra_data
        }
        with open(extra_log_file, 'a', encoding='utf-8') as f:
            f.write(json.dumps(extra_entry) + '\n')

def log_data_source_details(zip_code: str, source: str, raw_data: dict, processed_data: dict):
    """Log detailed information about data sources and processing"""
    timestamp = datetime.now().isoformat()

    log_entry = {
        'timestamp': timestamp,
        'zip_code': zip_code,
        'data_source': source,
        'raw_data': raw_data,
        'processed_data': processed_data
    }

    # Log to data sources file
    sources_log_file = os.path.join(LOGS_DIR, 'data_sources.jsonl')
    with open(sources_log_file, 'a', encoding='utf-8') as f:
        f.write(json.dumps(log_entry) + '\n')

def log_error(endpoint: str, zip_code: str, error: str, error_details: dict = None):
    """Log errors to separate error file"""
    timestamp = datetime.now().isoformat()

    error_entry = {
        'timestamp': timestamp,
        'endpoint': endpoint,
        'zip_code': zip_code,
        'error': error,
        'error_details': error_details or {}
    }

    error_log_file = os.path.join(LOGS_DIR, 'errors.jsonl')
    with open(error_log_file, 'a', encoding='utf-8') as f:
        f.write(json.dumps(error_entry) + '\n')

# API Keys and URLs from Environment Variables
EIA_API_KEY = os.getenv('EIA_API_KEY')
EIA_URL = os.getenv('EIA_URL')
ZIPPOPOTAM_URL = os.getenv('ZIPPOPOTAM_URL')
FCC_LOOKUP_URL = os.getenv('FCC_LOOKUP_URL')
CENSUS_API_KEY = os.getenv('CENSUS_API_KEY')
CENSUS_API_URL = os.getenv('CENSUS_API_URL')

# Vantage Score - now using local Excel file instead of API

# Validate required environment variables
required_env_vars = {
    'EIA_API_KEY': EIA_API_KEY,
    'EIA_URL': EIA_URL,
    'ZIPPOPOTAM_URL': ZIPPOPOTAM_URL,
    'FCC_LOOKUP_URL': FCC_LOOKUP_URL,
    'CENSUS_API_KEY': CENSUS_API_KEY,
    'CENSUS_API_URL': CENSUS_API_URL
}

missing_vars = [var for var, value in required_env_vars.items() if not value]
if missing_vars:
    logger.warning(f"Missing environment variables: {', '.join(missing_vars)}. Please check your .env file.")

# Global variable to cache Excel data
_vantage_data_cache = None

def load_vantage_data_from_excel():
    """Load Vantage Score data from local Excel file using openpyxl"""
    global _vantage_data_cache

    if _vantage_data_cache is not None:
        return _vantage_data_cache

    try:
        # Path to Excel file (in root directory)
        excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'new data.xlsx')

        if not os.path.exists(excel_path):
            logger.error(f"Excel file not found at: {excel_path}")
            return None

        # Read Excel file using openpyxl
        logger.info(f"Loading Vantage Score data from: {excel_path}")
        workbook = load_workbook(excel_path, read_only=True)
        worksheet = workbook.active

        # Get header row to find column indices
        headers = []
        for cell in worksheet[1]:
            headers.append(cell.value)

        # Find column indices for ZIP code and Vantage Score
        zip_col_idx = None
        vantage_col_idx = None
        city_col_idx = None
        state_col_idx = None

        for i, header in enumerate(headers):
            if header:
                header_lower = str(header).lower().strip()
                # Look for ZIP code column
                if header_lower in ['zip', 'zip_code', 'zipcode']:
                    zip_col_idx = i
                # Look for Vantage Score column
                elif 'vantage' in header_lower or 'score' in header_lower:
                    vantage_col_idx = i
                # Look for city column
                elif header_lower == 'city':
                    city_col_idx = i
                # Look for state column
                elif header_lower == 'state':
                    state_col_idx = i

        if zip_col_idx is None or vantage_col_idx is None:
            logger.error("Could not find ZIP code or Vantage Score columns in Excel file")
            return None

        # Convert to dictionary for faster lookups
        vantage_dict = {}

        # Read data rows (skip header row)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if len(row) > max(zip_col_idx, vantage_col_idx):
                zip_value = row[zip_col_idx]
                vantage_value = row[vantage_col_idx]

                if zip_value is not None and vantage_value is not None:
                    # Clean and format ZIP code
                    zip_code = str(zip_value).strip()
                    if len(zip_code) < 5:
                        zip_code = zip_code.zfill(5)  # Pad with leading zeros

                    # Get city and state if available
                    city = row[city_col_idx] if city_col_idx is not None and city_col_idx < len(row) else 'Unknown'
                    state = row[state_col_idx] if state_col_idx is not None and state_col_idx < len(row) else 'Unknown'

                    try:
                        vantage_dict[zip_code] = {
                            'vantage_score': float(vantage_value),
                            'city': str(city) if city else 'Unknown',
                            'state': str(state) if state else 'Unknown'
                        }
                    except (ValueError, TypeError):
                        # Skip rows with invalid vantage scores
                        continue

        workbook.close()
        _vantage_data_cache = vantage_dict
        logger.info(f"Loaded {len(vantage_dict)} Vantage Score records from Excel file")
        return vantage_dict

    except Exception as e:
        logger.error(f"Error loading Excel file: {e}")
        return None

# Gemini AI Configuration
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY')
if GEMINI_API_KEY:
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        logger.info("Gemini AI configured successfully")
    except Exception as e:
        logger.warning("Gemini AI configuration failed: %s", e)
else:
    logger.warning("GEMINI_API_KEY not found in environment variables. Please set it in .env file.")

def zip_to_location(zip_code: str):
    """Get location info from ZIP"""
    resp = requests.get(ZIPPOPOTAM_URL.format(zip=zip_code), timeout=10)
    resp.raise_for_status()
    data = resp.json()
    
    place = data['places'][0]
    lat, lng = float(place['latitude']), float(place['longitude'])
    state_code = place['state abbreviation']
    city = place['place name']
    
    # Get county
    params = {'latitude': lat, 'longitude': lng, 'format': 'json'}
    resp = requests.get(FCC_LOOKUP_URL, params=params, timeout=10)
    resp.raise_for_status()
    fcc_data = resp.json()
    
    county = fcc_data['County']['name'].replace(' County', '').lower().replace(' ', '-')
    state_slug = state_code.lower()
    
    logger.info("Location: %s, %s -> %s county", city, state_code, county)
    return county, state_slug, city, state_code

def try_findenergy_simple(county: str, state: str):
    """Simple attempt at findenergy.com"""
    try:
        url = f"https://findenergy.com/{state}/{county}-electricity/"
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        
        logger.info("Trying findenergy.com...")
        resp = requests.get(url, headers=headers, timeout=10)
        
        if resp.status_code == 200:
            text = BeautifulSoup(resp.text, 'html.parser').get_text()
            
            # Quick extraction
            bill_match = re.search(r'average.*?bill.*?\$([0-9,]+\.?[0-9]*)', text, re.IGNORECASE)
            usage_match = re.search(r'([0-9,]+)\s*kWh.*?per month', text, re.IGNORECASE)
            rate_match = re.search(r'([0-9]+\.?[0-9]*)\s*cents?\s*per\s*kWh', text, re.IGNORECASE)
            
            data = {}
            if bill_match:
                data['average_monthly_bill'] = float(bill_match.group(1).replace(',', ''))
            if usage_match:
                data['average_monthly_usage_kwh'] = float(usage_match.group(1).replace(',', ''))
            if rate_match:
                data['utility_rate_per_kwh'] = float(rate_match.group(1)) / 100
            
            if data:
                logger.info("FindEnergy data: %s", data)
                # Log raw scraped data for analysis
                raw_data = {
                    'url': url,
                    'bill_match': bill_match.group(0) if bill_match else None,
                    'usage_match': usage_match.group(0) if usage_match else None,
                    'rate_match': rate_match.group(0) if rate_match else None,
                    'response_length': len(resp.text)
                }
                return data, "findenergy.com", raw_data

            return None, None, None
                
    except Exception as e:
        logger.warning("FindEnergy failed: %s", e)

    return None, None, None

def get_eia_data(state_code: str):
    """Get real-time data from EIA as fallback"""
    try:
        params = {
            'api_key': EIA_API_KEY,
            'frequency': 'monthly',
            'data[0]': 'sales',
            'data[1]': 'revenue', 
            'data[2]': 'customers',
            'facets[stateid][]': state_code,
            'facets[sectorid][]': 'RES',
            'sort[0][column]': 'period',
            'sort[0][direction]': 'desc',
            'offset': 0,
            'length': 1
        }
        
        logger.info("Getting EIA data for %s...", state_code)
        resp = requests.get(EIA_URL, params=params, timeout=15)
        resp.raise_for_status()
        
        data = resp.json()['response']['data'][0]
        
        total_kwh = float(data['sales']) * 1000000  # Million kWh to kWh
        total_revenue = float(data['revenue']) * 1000000  # Million $ to $
        total_customers = float(data['customers'])
        
        result = {
            'average_monthly_usage_kwh': round(total_kwh / total_customers),
            'utility_rate_per_kwh': round(total_revenue / total_kwh, 4),
            'average_monthly_bill': round(total_revenue / total_customers, 2),
            'period': data['period']
        }
        
        logger.info("EIA data: %s", result)
        return result, f"EIA (period: {data['period']})"
        
    except Exception as e:
        logger.error("EIA failed: %s", e)
        return None, None

def try_alternative_sources(state_code: str):
    """Try other real-time sources"""
    
    # Source 1: Try electricityrates.com
    try:
        url = f"https://www.electricityrates.com/electricity-rates/{state_code.lower()}/"
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        
        logger.info("Trying electricityrates.com...")
        resp = requests.get(url, headers=headers, timeout=10)
        
        if resp.status_code == 200:
            text = BeautifulSoup(resp.text, 'html.parser').get_text()
            
            # Look for rate data
            rate_match = re.search(r'([0-9]+\.?[0-9]*)\s*cents?\s*per\s*kWh', text, re.IGNORECASE)
            usage_match = re.search(r'average.*?([0-9,]+)\s*kWh', text, re.IGNORECASE)
            
            if rate_match:
                rate = float(rate_match.group(1)) / 100
                usage = 900  # Default usage
                if usage_match:
                    usage = float(usage_match.group(1).replace(',', ''))
                
                data = {
                    'utility_rate_per_kwh': rate,
                    'average_monthly_usage_kwh': usage,
                    'average_monthly_bill': round(rate * usage, 2)
                }
                
                logger.info("ElectricityRates data: %s", data)
                return data, "electricityrates.com"
                
    except Exception as e:
        logger.warning("ElectricityRates failed: %s", e)
    
    # Source 2: Try saveonenergy.com
    try:
        url = f"https://www.saveonenergy.com/electricity-rates/{state_code.lower()}/"
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        
        logger.info("Trying saveonenergy.com...")
        resp = requests.get(url, headers=headers, timeout=10)
        
        if resp.status_code == 200:
            text = BeautifulSoup(resp.text, 'html.parser').get_text()
            
            rate_match = re.search(r'([0-9]+\.?[0-9]*)\s*cents?\s*per\s*kWh', text, re.IGNORECASE)
            if rate_match:
                rate = float(rate_match.group(1)) / 100
                usage = 900  # Default
                
                data = {
                    'utility_rate_per_kwh': rate,
                    'average_monthly_usage_kwh': usage,
                    'average_monthly_bill': round(rate * usage, 2)
                }
                
                logger.info("SaveOnEnergy data: %s", data)
                return data, "saveonenergy.com"
                
    except Exception as e:
        logger.warning("SaveOnEnergy failed: %s", e)
    
    return None, None

def get_census_demographics(zip_code: str):
    """Get race and income data from Census API"""
    try:
        variables = [
            "NAME",
            "B02001_001E",  # Total population
            "B02001_002E",  # White alone
            "B02001_003E",  # Black alone
            "B02001_004E",  # American Indian/Alaska Native alone
            "B02001_005E",  # Asian alone
            "B02001_006E",  # Native Hawaiian/Pacific Islander alone
            "B02001_007E",  # Some other race alone
            "B02001_008E",  # Two or more races
            "B19013_001E",  # Median household income
        ]
        
        url = f"{CENSUS_API_URL}?get={','.join(variables)}&for=zip%20code%20tabulation%20area:{zip_code}&key={CENSUS_API_KEY}"
        
        logger.info("Fetching Census data for ZIP %s", zip_code)
        resp = requests.get(url, timeout=10)
        resp.raise_for_status()
        
        data = resp.json()
        headers = data[0]
        values = data[1]
        result = dict(zip(headers, values))
        
        # Calculate percentages
        total_pop = int(result["B02001_001E"])
        demographics = {
            'total_population': total_pop,
            'median_household_income': int(result["B19013_001E"]),
            'race_breakdown': {
                'white': int(result["B02001_002E"]),
                'black': int(result["B02001_003E"]),
                'asian': int(result["B02001_005E"]),
                'native_american': int(result["B02001_004E"]),
                'pacific_islander': int(result["B02001_006E"]),
                'other': int(result["B02001_007E"]),
                'mixed': int(result["B02001_008E"])
            }
        }
        
        if total_pop > 0:
            demographics['race_percentages'] = {
                'white': round((demographics['race_breakdown']['white'] / total_pop) * 100, 1),
                'black': round((demographics['race_breakdown']['black'] / total_pop) * 100, 1),
                'asian': round((demographics['race_breakdown']['asian'] / total_pop) * 100, 1),
                'native_american': round((demographics['race_breakdown']['native_american'] / total_pop) * 100, 1),
                'pacific_islander': round((demographics['race_breakdown']['pacific_islander'] / total_pop) * 100, 1),
                'other': round((demographics['race_breakdown']['other'] / total_pop) * 100, 1),
                'mixed': round((demographics['race_breakdown']['mixed'] / total_pop) * 100, 1)
            }
        
        logger.info("Successfully fetched Census data")
        return demographics
        
    except Exception as e:
        logger.warning("Failed to get Census data: %s", e)
        return None

@app.route('/electricity-data')
def electricity_data():
    zip_code = request.args.get('zip', '').strip()

    if not zip_code.isdigit() or len(zip_code) != 5:
        error_msg = 'Invalid ZIP code'
        log_error('electricity-data', zip_code, error_msg)
        return jsonify({'error': error_msg}), 400

    try:
        # Get location info
        county, state, city, state_code = zip_to_location(zip_code)
        location_data = {
            'county': county,
            'state': state,
            'city': city,
            'state_code': state_code
        }

        # Try findenergy.com first
        data, source, raw_data = try_findenergy_simple(county, state)

        # Fallback to EIA
        if not data:
            data, source = get_eia_data(state_code)
            raw_data = None

        # Fallback to alternative sources
        if not data:
            data, source = try_alternative_sources(state_code)
            raw_data = None

        if data:
            response_data = {
                'zip_code': zip_code,
                'city': city,
                'state': state_code,
                'data_source': source,
                **data
            }

            # Log the successful request with extra location data
            extra_data = {
                'location_details': location_data,
                'data_source_used': source,
                'raw_scraped_data': raw_data,
                'request_ip': request.remote_addr,
                'user_agent': request.headers.get('User-Agent', 'Unknown')
            }

            log_api_request('electricity-data', zip_code, response_data, extra_data)
            log_data_source_details(zip_code, source, {'processed': data, 'raw': raw_data}, response_data)

            return jsonify(response_data)
        else:
            error_msg = 'No data available'
            log_error('electricity-data', zip_code, error_msg, {'location_data': location_data})
            return jsonify({'error': error_msg}), 404

    except Exception as e:
        error_msg = str(e)
        logger.error("Error: %s", e)
        log_error('electricity-data', zip_code, error_msg, {'exception_type': type(e).__name__})
        return jsonify({'error': error_msg}), 500

@app.route('/demographic-data')
def demographic_data():
    zip_code = request.args.get('zip', '').strip()

    if not zip_code.isdigit() or len(zip_code) != 5:
        error_msg = 'Invalid ZIP code'
        log_error('demographic-data', zip_code, error_msg)
        return jsonify({'error': error_msg}), 400

    try:
        demographics = get_census_demographics(zip_code)
        if demographics:
            # Get city/state info for nicer response
            _, _, city, state_code = zip_to_location(zip_code)

            response_data = {
                'zip_code': zip_code,
                'city': city,
                'state': state_code,
                'data_source': 'U.S. Census Bureau ACS 5-year estimates (2021)',
                **demographics
            }

            # Log the successful request with detailed demographic data
            extra_data = {
                'census_api_used': True,
                'total_population': demographics.get('total_population'),
                'median_income': demographics.get('median_household_income'),
                'race_diversity_score': calculate_diversity_score(demographics.get('race_percentages', {})),
                'request_ip': request.remote_addr,
                'user_agent': request.headers.get('User-Agent', 'Unknown')
            }

            log_api_request('demographic-data', zip_code, response_data, extra_data)
            log_data_source_details(zip_code, 'U.S. Census Bureau', demographics, response_data)

            return jsonify(response_data)
        else:
            error_msg = 'No demographic data available'
            log_error('demographic-data', zip_code, error_msg)
            return jsonify({'error': error_msg}), 404

    except Exception as e:
        error_msg = str(e)
        logger.error("Error: %s", e)
        log_error('demographic-data', zip_code, error_msg, {'exception_type': type(e).__name__})
        return jsonify({'error': error_msg}), 500

def calculate_diversity_score(race_percentages: dict) -> float:
    """Calculate a simple diversity score based on race percentages"""
    if not race_percentages:
        return 0.0

    # Simpson's Diversity Index (1 - sum of squares of proportions)
    total = 0.0
    for percentage in race_percentages.values():
        proportion = percentage / 100.0
        total += proportion * proportion

    return round(1.0 - total, 3)

def get_vantage_score(zip_code: str):
    """Get average Vantage Score for ZIP code from local Excel file"""
    try:
        # Load data from local Excel file
        vantage_data = load_vantage_data_from_excel()

        if not vantage_data:
            logger.error("Failed to load Vantage Score data from Excel file")
            return None

        # Look up ZIP code in local data
        if zip_code in vantage_data:
            record = vantage_data[zip_code]
            logger.info(f"Vantage Score found for ZIP {zip_code}: {record['vantage_score']}")
            return {
                'zip_code': zip_code,
                'vantage_score': record['vantage_score'],
                'source': 'Local Excel File',
                'city': record.get('city', 'Unknown'),
                'state': record.get('state', 'Unknown')
            }
        else:
            logger.warning(f"No Vantage Score found for ZIP {zip_code} in local data")
            return None

    except Exception as e:
        logger.error(f"Vantage Score local lookup error: {e}")
        return None

def calculate_solar_qualification_with_gemini(zip_code: str, monthly_bill: float, credit_band: str, roof_size: float):
    """Use Gemini AI to calculate solar loan qualification"""
    try:
        # Get electricity and demographic data
        county, state, city, state_code = zip_to_location(zip_code)

        # Get electricity data
        electricity_data, source, _ = try_findenergy_simple(county, state)
        if not electricity_data:
            electricity_data, source = get_eia_data(state_code)
        if not electricity_data:
            electricity_data, source = try_alternative_sources(state_code)

        # Skip demographic data to keep analysis simple
        demographics = None

        # Prepare data for Gemini
        context_data = {
            'location': {
                'zip_code': zip_code,
                'city': city,
                'state': state_code,
                'county': county
            },
            'electricity': electricity_data or {
                'average_monthly_bill': monthly_bill,
                'utility_rate_per_kwh': 0.15,  # Default rate
                'average_monthly_usage_kwh': monthly_bill / 0.15
            },
            'demographics': demographics,
            'user_input': {
                'monthly_bill': monthly_bill,
                'credit_band': credit_band,
                'roof_size': roof_size
            }
        }

        # Create Gemini prompt
        prompt = f"""
You are an expert solar loan qualification analyst. Based on the following data, calculate and determine solar loan qualification.

LOCATION DATA:
- ZIP Code: {zip_code}
- City: {city}, {state_code}
- County: {county}

ELECTRICITY DATA:
- User's Exact Monthly Bill: ${monthly_bill:.0f}
- Local Average Bill: ${electricity_data.get('average_monthly_bill', 'N/A') if electricity_data else 'N/A'}
- Utility Rate: ${electricity_data.get('utility_rate_per_kwh', 0.15) if electricity_data else 0.15}/kWh
- User's Estimated Usage: {monthly_bill/0.15:.0f} kWh/month (based on their exact bill amount)

ANALYSIS FOCUS:
- Focus on user's individual profile and electricity usage
- Do not consider area demographics or median income data
- Base qualification solely on user's credit band, bill amount, and roof size

USER PROFILE:
- Monthly Electric Bill: ${monthly_bill}
- Credit Band: {credit_band}
- Available Roof Size: {roof_size} sq ft
- Maximum System Capacity: {roof_size / 250:.1f} kW (based on roof size constraint)

SOLAR INDUSTRY STANDARDS:
- Solar panels: ~400W each, ~25 sq ft per panel (including spacing)
- System cost: ~$2.75/watt installed
- Federal tax credit: 30%
- Typical sun hours: 4-6 hours/day depending on location
- System efficiency: ~85% (including inverter losses)
- Panel degradation: 0.5% per year
- Roof space requirement: ~250 sq ft per kW (realistic spacing)
- Loan terms by credit:
  * Excellent (750+): 3.99% APR, 25 years, 0% down
  * Good (700-749): 5.99% APR, 20 years, 0% down
  * Fair (650-699): 8.99% APR, 15 years, 10% down
  * Poor (<650): 12.99% APR, 10 years, 20% down

CALCULATE AND PROVIDE:
1. Recommended system size (kW) - MUST NOT exceed roof capacity limit
2. Total system cost (before and after incentives)
3. 25-year lifetime savings
4. Qualification status: "approved", "borderline", or "not_qualified"
5. Brief explanation (2-3 sentences)

IMPORTANT: System size CANNOT exceed {roof_size / 250:.1f} kW due to roof space limitations.

QUALIFICATION RULES:
- Excellent/Good Credit: Always "approved"
- Fair Credit: Always "borderline" (requires review)
- Poor Credit: Always "not_qualified"

Consider factors like:
- Roof space constraints (system size MUST fit available roof area)
- Balance between electricity usage needs and roof capacity
- Local electricity rates and solar potential
- Credit band for qualification status
- Realistic panel spacing and installation requirements

Respond in this exact JSON format:
{{
    "status": "approved|borderline|not_qualified",
    "system_size_kw": 0.0,
    "total_cost": 0.0,
    "net_cost_after_incentives": 0.0,
    "lifetime_savings": 0.0,
    "explanation": "Brief explanation here",
    "loan_terms": {{
        "apr": 0.0,
        "term_years": 0,
        "down_payment_percent": 0
    }},
    "calculations": {{
        "monthly_kwh_usage": 0.0,
        "system_annual_production": 0.0
    }}
}}
"""

        # Call Gemini API
        model = genai.GenerativeModel('gemini-1.5-flash')
        response = model.generate_content(prompt)

        # Parse JSON response
        result_text = response.text.strip()
        if result_text.startswith('```json'):
            result_text = result_text[7:-3]
        elif result_text.startswith('```'):
            result_text = result_text[3:-3]

        result = json.loads(result_text)

        # Log the Gemini calculation
        log_gemini_calculation(zip_code, context_data, result)

        return result

    except Exception as e:
        logger.error("Gemini calculation failed: %s", e)
        # Fallback to simple calculation
        return fallback_calculation(monthly_bill, credit_band, roof_size)

def log_gemini_calculation(zip_code: str, input_data: dict, result: dict):
    """Log Gemini AI calculations for analysis"""
    timestamp = datetime.now().isoformat()

    log_entry = {
        'timestamp': timestamp,
        'zip_code': zip_code,
        'ai_model': 'gemini-1.5-flash',
        'input_data': input_data,
        'ai_result': result
    }

    gemini_log_file = os.path.join(LOGS_DIR, 'gemini_calculations.jsonl')
    with open(gemini_log_file, 'a', encoding='utf-8') as f:
        f.write(json.dumps(log_entry) + '\n')

def fallback_calculation(monthly_bill: float, credit_band: str, roof_size: float):
    """Simple fallback calculation if Gemini fails"""
    # Basic calculations using exact bill amount
    estimated_usage = monthly_bill / 0.15  # Assume 15¢/kWh
    usage_based_size = (estimated_usage * 12) / (365 * 5 * 0.85) / 1000  # 5 sun hours, 85% efficiency

    # Roof size constraints - more realistic sizing based on available space
    # Assume 250 sq ft per kW for better spacing and realistic installation
    max_roof_capacity = roof_size / 250

    # System size should be constrained by both usage and roof space
    # But give more weight to roof size to show meaningful differences
    system_size = min(usage_based_size, max_roof_capacity)

    # Ensure minimum viable system size
    system_size = max(2.0, system_size)

    gross_cost = system_size * 1000 * 2.75
    net_cost = gross_cost * 0.7  # 30% federal credit

    # Loan terms by credit
    loan_terms = {
        'Excellent': {'apr': 3.99, 'years': 25, 'down': 0},
        'Good': {'apr': 5.99, 'years': 20, 'down': 0},
        'Fair': {'apr': 8.99, 'years': 15, 'down': 10},
        'Poor': {'apr': 12.99, 'years': 10, 'down': 20}
    }

    terms = loan_terms.get(credit_band, loan_terms['Fair'])

    # Remove monthly payment calculation - not needed

    # Determine status based on credit band only
    # Credit band based qualification logic
    if credit_band in ['Excellent', 'Good']:
        # Excellent and Good credit get approved
        status = 'approved'
    elif credit_band == 'Fair':
        # Fair credit gets borderline status for review
        status = 'borderline'
    else:  # Poor credit
        # Poor credit is not qualified
        status = 'not_qualified'

    return {
        'status': status,
        'system_size_kw': round(system_size, 2),
        'total_cost': round(gross_cost, 2),
        'net_cost_after_incentives': round(net_cost, 2),
        'lifetime_savings': round(monthly_bill * 12 * 25 - net_cost, 2),
        'explanation': f"Based on your ${monthly_bill} monthly bill and {credit_band} credit, this {system_size:.1f}kW system is recommended.",
        'loan_terms': {
            'apr': terms['apr'],
            'term_years': terms['years'],
            'down_payment_percent': terms['down']
        },
        'calculations': {
            'monthly_kwh_usage': round(estimated_usage, 0),
            'system_annual_production': round(system_size * 365 * 5 * 0.85, 0)
        }
    }

@app.route('/logs/summary')
def logs_summary():
    """Get a summary of all logged data"""
    try:
        summary = {
            'total_requests': 0,
            'unique_zip_codes': set(),
            'endpoints_used': {},
            'data_sources_used': {},
            'errors_count': 0,
            'recent_requests': []
        }

        # Read API requests log
        api_log_file = os.path.join(LOGS_DIR, 'api_requests.jsonl')
        if os.path.exists(api_log_file):
            with open(api_log_file, 'r', encoding='utf-8') as f:
                for line in f:
                    try:
                        entry = json.loads(line.strip())
                        summary['total_requests'] += 1
                        summary['unique_zip_codes'].add(entry['zip_code'])

                        endpoint = entry['endpoint']
                        summary['endpoints_used'][endpoint] = summary['endpoints_used'].get(endpoint, 0) + 1

                        if 'data_source' in entry['response_data']:
                            source = entry['response_data']['data_source']
                            summary['data_sources_used'][source] = summary['data_sources_used'].get(source, 0) + 1

                        # Keep last 10 requests
                        if len(summary['recent_requests']) < 10:
                            summary['recent_requests'].append({
                                'timestamp': entry['timestamp'],
                                'endpoint': entry['endpoint'],
                                'zip_code': entry['zip_code']
                            })
                    except json.JSONDecodeError:
                        continue

        # Read errors log
        error_log_file = os.path.join(LOGS_DIR, 'errors.jsonl')
        if os.path.exists(error_log_file):
            with open(error_log_file, 'r', encoding='utf-8') as f:
                for line in f:
                    try:
                        json.loads(line.strip())
                        summary['errors_count'] += 1
                    except json.JSONDecodeError:
                        continue

        # Convert set to list for JSON serialization
        summary['unique_zip_codes'] = len(summary['unique_zip_codes'])

        return jsonify(summary)

    except Exception as e:
        logger.error("Error reading logs: %s", e)
        return jsonify({'error': str(e)}), 500

@app.route('/logs/<log_type>')
def get_logs(log_type):
    """Get specific log file contents"""
    try:
        valid_log_types = [
            'api_requests', 'electricity_data_data', 'demographic_data_data',
            'electricity_data_extra_data', 'demographic_data_extra_data',
            'data_sources', 'errors'
        ]

        if log_type not in valid_log_types:
            return jsonify({'error': 'Invalid log type'}), 400

        log_file = os.path.join(LOGS_DIR, f'{log_type}.jsonl')

        if not os.path.exists(log_file):
            return jsonify({'logs': [], 'message': 'Log file does not exist yet'})

        logs = []
        with open(log_file, 'r', encoding='utf-8') as f:
            for line in f:
                try:
                    logs.append(json.loads(line.strip()))
                except json.JSONDecodeError:
                    continue

        # Return most recent 100 entries
        return jsonify({'logs': logs[-100:], 'total_entries': len(logs)})

    except Exception as e:
        logger.error("Error reading log file: %s", e)
        return jsonify({'error': str(e)}), 500

@app.route('/api/check-qualification', methods=['POST'])
def check_qualification():
    """Solar loan qualification endpoint using Gemini AI"""
    try:
        data = request.get_json()

        # Validate input
        required_fields = ['zipCode', 'electricBill', 'creditBand', 'roofSize']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400

        zip_code = str(data['zipCode']).strip()
        monthly_bill = float(data['electricBill'])
        credit_band = str(data['creditBand'])
        roof_size = float(data['roofSize'])

        # Validate ZIP code
        if not zip_code.isdigit() or len(zip_code) != 5:
            return jsonify({'error': 'Invalid ZIP code format'}), 400

        # Validate ranges
        if monthly_bill < 50 or monthly_bill > 500:
            return jsonify({'error': 'Electric bill must be between $50 and $500'}), 400

        if roof_size <= 0 or roof_size > 50000:
            return jsonify({'error': 'Invalid roof size'}), 400

        if credit_band not in ['Excellent', 'Good', 'Fair', 'Poor']:
            return jsonify({'error': 'Invalid credit band'}), 400

        # Calculate qualification using Gemini AI
        result = calculate_solar_qualification_with_gemini(
            zip_code, monthly_bill, credit_band, roof_size
        )

        # Add location information to the result
        try:
            _, _, city, state_code = zip_to_location(zip_code)
            result['location'] = {
                'city': city,
                'state': state_code,
                'zip_code': zip_code
            }
        except Exception as e:
            logger.warning(f"Could not get location info for {zip_code}: {e}")
            result['location'] = {
                'city': 'Unknown',
                'state': 'Unknown',
                'zip_code': zip_code
            }

        # Log the qualification request
        log_api_request('check-qualification', zip_code, result, {
            'input_data': data,
            'ai_powered': True,
            'request_ip': request.remote_addr,
            'user_agent': request.headers.get('User-Agent', 'Unknown')
        })

        return jsonify(result)

    except ValueError as e:
        error_msg = f'Invalid input data: {str(e)}'
        log_error('check-qualification', data.get('zipCode', 'unknown'), error_msg)
        return jsonify({'error': error_msg}), 400

    except Exception as e:
        error_msg = str(e)
        logger.error("Qualification error: %s", e)
        log_error('check-qualification', data.get('zipCode', 'unknown'), error_msg, {
            'exception_type': type(e).__name__,
            'input_data': data
        })
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/vantage-score')
def vantage_score():
    """Get Vantage Score for ZIP code"""
    zip_code = request.args.get('zip', '').strip()

    if not zip_code.isdigit() or len(zip_code) != 5:
        error_msg = 'Invalid ZIP code'
        log_error('vantage-score', zip_code, error_msg)
        return jsonify({'error': error_msg}), 400

    try:
        vantage_data = get_vantage_score(zip_code)

        if vantage_data:
            # Log the successful request
            extra_data = {
                'local_excel_used': True,
                'data_source': vantage_data.get('source', 'Local Excel File'),
                'request_ip': request.remote_addr,
                'user_agent': request.headers.get('User-Agent', 'Unknown')
            }

            log_api_request('vantage-score', zip_code, vantage_data, extra_data)

            return jsonify(vantage_data)
        else:
            error_msg = 'No Vantage Score data available for this ZIP code'
            log_error('vantage-score', zip_code, error_msg)
            return jsonify({'error': error_msg}), 404

    except Exception as e:
        error_msg = str(e)
        logger.error("Vantage Score error: %s", e)
        log_error('vantage-score', zip_code, error_msg, {'exception_type': type(e).__name__})
        return jsonify({'error': 'Internal server error'}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5500))
    host = os.environ.get('HOST', '0.0.0.0')
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(debug=debug, host=host, port=port)